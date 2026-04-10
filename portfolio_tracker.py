"""
Stock Portfolio Tracker
Fetches live stock data via yfinance and generates a formatted Excel report.
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import sys

# ── Portfolio Definition ────────────────────────────────────────────────────
# Edit this list: (ticker, shares_owned, avg_cost_per_share)
PORTFOLIO = [
    ("AAPL",  10,  150.00),
    ("MSFT",   5,  280.00),
    ("GOOGL",  3, 2800.00),
    ("AMZN",   8,  170.00),
    ("JPM",   15,  185.00),
    ("BAC",   20,   38.00),
    ("GS",     4,  410.00),
    ("NVDA",   6,  500.00),
]

OUTPUT_FILE = "portfolio_report.xlsx"

# ── Color Palette ───────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_BLUE    = "1B3A6B"
ACCENT_BLUE = "2563EB"
LIGHT_BLUE  = "DBEAFE"
GREEN       = "166534"
GREEN_BG    = "DCFCE7"
RED         = "991B1B"
RED_BG      = "FEE2E2"
GOLD        = "D97706"
GOLD_BG     = "FEF3C7"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F8FAFC"
MED_GRAY    = "E2E8F0"
DARK_TEXT   = "1E293B"


def fetch_data(portfolio):
    tickers = [t for t, _, _ in portfolio]
    raw = yf.download(tickers, period="1y", auto_adjust=True, progress=False)

    records = []
    for ticker, shares, avg_cost in portfolio:
        try:
            info   = yf.Ticker(ticker).info
            hist   = raw["Close"][ticker].dropna()

            price      = float(hist.iloc[-1])
            prev_close = float(hist.iloc[-2]) if len(hist) > 1 else price
            week_ago   = float(hist.iloc[-6]) if len(hist) > 5 else price
            month_ago  = float(hist.iloc[-22]) if len(hist) > 21 else price
            year_ago   = float(hist.iloc[0])

            records.append({
                "Ticker":        ticker,
                "Company":       info.get("shortName", ticker),
                "Sector":        info.get("sector", "N/A"),
                "Shares":        shares,
                "Avg Cost":      avg_cost,
                "Current Price": price,
                "Prev Close":    prev_close,
                "1W Ago":        week_ago,
                "1M Ago":        month_ago,
                "1Y Ago":        year_ago,
                "Market Cap":    info.get("marketCap", 0),
                "PE Ratio":      info.get("trailingPE", None),
                "Div Yield":     info.get("dividendYield", None),
                "Beta":          info.get("beta", None),
                "52W High":      info.get("fiftyTwoWeekHigh", None),
                "52W Low":       info.get("fiftyTwoWeekLow", None),
            })
        except Exception as e:
            print(f"  Warning: could not fetch {ticker}: {e}")

    return pd.DataFrame(records)


def thin_border():
    s = Side(style="thin", color=MED_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)


def body_font(size=10, bold=False, color=DARK_TEXT):
    return Font(name="Arial", size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def right():
    return Alignment(horizontal="right", vertical="center")


def build_workbook(df):
    wb = Workbook()

    _build_summary(wb, df)
    _build_holdings(wb, df)
    _build_performance(wb, df)
    _build_fundamentals(wb, df)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved → {OUTPUT_FILE}")


# ── Sheet 1: Summary Dashboard ───────────────────────────────────────────────
def _build_summary(wb, df):
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    # Title block
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:I2")
    t = ws["B2"]
    t.value         = "STOCK PORTFOLIO TRACKER"
    t.font          = Font(name="Arial", size=22, bold=True, color=WHITE)
    t.fill          = fill(DARK_NAVY)
    t.alignment     = center()

    ws.merge_cells("B3:I3")
    sub = ws["B3"]
    sub.value       = f"Generated  {datetime.now().strftime('%B %d, %Y  |  %I:%M %p')}"
    sub.font        = Font(name="Arial", size=10, color="94A3B8")
    sub.fill        = fill(DARK_NAVY)
    sub.alignment   = center()
    ws.row_dimensions[3].height = 20

    # KPI row
    total_invested  = (df["Shares"] * df["Avg Cost"]).sum()
    total_value     = (df["Shares"] * df["Current Price"]).sum()
    total_gain      = total_value - total_invested
    total_pct       = total_gain / total_invested
    daily_chg       = (df["Shares"] * (df["Current Price"] - df["Prev Close"])).sum()
    daily_pct       = daily_chg / (df["Shares"] * df["Prev Close"]).sum()

    kpis = [
        ("Total Value",       f"${total_value:,.2f}",  None),
        ("Total Invested",    f"${total_invested:,.2f}", None),
        ("Total Gain / Loss", f"${total_gain:+,.2f}",  total_gain),
        ("Return %",          f"{total_pct:+.2%}",     total_pct),
        ("Day's Change",      f"${daily_chg:+,.2f}",   daily_chg),
        ("Positions",         str(len(df)),              None),
    ]

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 55
    ws.row_dimensions[7].height = 20

    kpi_cols = ["B","C","D","E","F","G","H","I"]
    col_pairs = [("B","C"), ("C","D"), ("D","E"), ("E","F"), ("F","G"), ("G","H")]

    for i, (label, value, signal) in enumerate(kpis):
        col = kpi_cols[i + 1]
        bg  = DARK_NAVY
        if signal is not None:
            bg = GREEN_BG if signal >= 0 else RED_BG
        val_color = GREEN if (signal is not None and signal >= 0) else \
                    RED   if (signal is not None and signal <  0) else MID_BLUE

        ws.merge_cells(f"{col}6:{col}6")
        cell_label = ws[f"{col}6"]
        cell_label.value     = f"{label}\n{value}"
        cell_label.font      = Font(name="Arial", size=11, bold=True, color=val_color)
        cell_label.fill      = fill(bg if signal is None else bg)
        cell_label.alignment = center()
        cell_label.border    = thin_border()
        ws.column_dimensions[col].width = 18

    # Allocation chart placeholder header
    ws.row_dimensions[9].height = 24
    ws.merge_cells("B9:I9")
    h = ws["B9"]
    h.value     = "PORTFOLIO ALLOCATION"
    h.font      = Font(name="Arial", size=12, bold=True, color=WHITE)
    h.fill      = fill(MID_BLUE)
    h.alignment = center()

    # Allocation table
    headers = ["Ticker", "Company", "Sector", "Shares", "Mkt Value ($)", "Weight (%)"]
    header_cols = ["B","C","D","E","F","G"]
    ws.row_dimensions[10].height = 22

    for col, hdr in zip(header_cols, headers):
        c = ws[f"{col}10"]
        c.value     = hdr
        c.font      = header_font(size=10)
        c.fill      = fill(ACCENT_BLUE)
        c.alignment = center()
        c.border    = thin_border()

    total_val = (df["Shares"] * df["Current Price"]).sum()
    for i, row in df.iterrows():
        r     = 11 + i
        val   = row["Shares"] * row["Current Price"]
        data  = [
            row["Ticker"],
            row["Company"],
            row["Sector"],
            row["Shares"],
            val,
            val / total_val,
        ]
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        for col, v in zip(header_cols, data):
            c = ws[f"{col}{r}"]
            c.value  = v
            c.font   = body_font()
            c.fill   = fill(bg)
            c.border = thin_border()
            ws.row_dimensions[r].height = 18
            if col in ("E",):
                c.number_format = '$#,##0.00'
                c.alignment = right()
            elif col in ("F",):
                c.number_format = '0.00%'
                c.alignment = right()
            elif col in ("D",):
                c.number_format = '#,##0'
                c.alignment = right()
            else:
                c.alignment = Alignment(vertical="center")

    # Bar chart: portfolio value by ticker
    last_data_row = 10 + len(df)
    chart = BarChart()
    chart.type          = "col"
    chart.title         = "Market Value by Holding"
    chart.y_axis.title  = "Value ($)"
    chart.x_axis.title  = "Ticker"
    chart.style         = 10
    chart.height        = 10
    chart.width         = 22
    chart.grouping      = "clustered"

    data_ref   = Reference(ws, min_col=6, min_row=10, max_row=last_data_row)
    cats_ref   = Reference(ws, min_col=2, min_row=11, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = ACCENT_BLUE

    ws.add_chart(chart, "B" + str(last_data_row + 2))


# ── Sheet 2: Holdings Detail ──────────────────────────────────────────────────
def _build_holdings(wb, df):
    ws = wb.create_sheet("📋 Holdings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Ticker",         12, "str"),
        ("Company",        28, "str"),
        ("Sector",         22, "str"),
        ("Shares",         10, "int"),
        ("Avg Cost",       13, "usd"),
        ("Current Price",  14, "usd"),
        ("Cost Basis",     14, "usd_formula"),
        ("Market Value",   14, "usd_formula"),
        ("Unrealized P&L", 15, "usd_formula"),
        ("Return %",       12, "pct_formula"),
        ("Day Change $",   14, "usd_formula"),
        ("Day Change %",   12, "pct_formula"),
    ]

    # Header
    ws.row_dimensions[1].height = 28
    for ci, (hdr, width, _) in enumerate(columns, 1):
        col = get_column_letter(ci)
        c = ws.cell(1, ci, hdr)
        c.font      = header_font()
        c.fill      = fill(DARK_NAVY)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[col].width = width

    for i, row in df.iterrows():
        r  = i + 2
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        # Columns E=5 Avg Cost, F=6 Current Price, D=4 Shares
        # Cost Basis  = D*E   (col G=7)
        # Market Val  = D*F   (col H=8)
        # Unrealized  = H-G   (col I=9)
        # Return %    = I/G   (col J=10)
        # Day Chg $   = D*(F-PrevClose) — prev close stored in col F for formula
        # We'll place PrevClose in a hidden helper or just compute in Python for Day Change

        prev_close = row["Prev Close"]

        raw_values = [
            row["Ticker"],
            row["Company"],
            row["Sector"],
            row["Shares"],
            row["Avg Cost"],
            row["Current Price"],
        ]
        for ci, v in enumerate(raw_values, 1):
            c = ws.cell(r, ci, v)
            c.font   = body_font()
            c.fill   = fill(bg)
            c.border = thin_border()
            if ci in (4,):
                c.number_format = '#,##0'
                c.alignment = right()
            elif ci in (5, 6):
                c.number_format = '$#,##0.00'
                c.alignment = right()
            else:
                c.alignment = Alignment(vertical="center")

        # Formulas
        D = f"D{r}"
        E = f"E{r}"
        F = f"F{r}"

        formula_cells = [
            (7,  f"={D}*{E}",                  '$#,##0.00', None),
            (8,  f"={D}*{F}",                  '$#,##0.00', None),
            (9,  f"=H{r}-G{r}",               '$#,##0.00;($#,##0.00);"-"', row["Current Price"] - row["Avg Cost"]),
            (10, f"=IFERROR(I{r}/G{r},0)",    '0.00%;(0.00%);"-"',         row["Current Price"] / row["Avg Cost"] - 1),
            (11, f"={D}*({F}-{prev_close})",  '$#,##0.00;($#,##0.00);"-"', row["Current Price"] - prev_close),
            (12, f"=IFERROR(K{r}/(D{r}*{prev_close}),0)", '0.00%;(0.00%);"-"', (row["Current Price"] - prev_close) / prev_close),
        ]

        for ci, formula, fmt, signal in formula_cells:
            c = ws.cell(r, ci, formula)
            c.number_format = fmt
            c.fill          = fill(bg)
            c.border        = thin_border()
            c.alignment     = right()
            if signal is not None:
                c.font = Font(name="Arial", size=10,
                              color=GREEN if signal >= 0 else RED)
            else:
                c.font = body_font()

    # Totals row
    tr = len(df) + 2
    ws.row_dimensions[tr].height = 22
    for ci in range(1, 13):
        c = ws.cell(tr, ci)
        c.fill   = fill(MID_BLUE)
        c.border = thin_border()

    ws.cell(tr, 1, "TOTAL").font      = header_font()
    ws.cell(tr, 1).alignment          = center()
    ws.cell(tr, 7, f"=SUM(G2:G{tr-1})").number_format = '$#,##0.00'
    ws.cell(tr, 7).font = header_font()
    ws.cell(tr, 7).alignment = right()
    ws.cell(tr, 8, f"=SUM(H2:H{tr-1})").number_format = '$#,##0.00'
    ws.cell(tr, 8).font = header_font()
    ws.cell(tr, 8).alignment = right()
    ws.cell(tr, 9, f"=SUM(I2:I{tr-1})").number_format = '$#,##0.00;($#,##0.00);"-"'
    ws.cell(tr, 9).font = header_font()
    ws.cell(tr, 9).alignment = right()
    ws.cell(tr, 10, f"=IFERROR(I{tr}/G{tr},0)").number_format = '0.00%;(0.00%);"-"'
    ws.cell(tr, 10).font = header_font()
    ws.cell(tr, 10).alignment = right()
    ws.cell(tr, 11, f"=SUM(K2:K{tr-1})").number_format = '$#,##0.00;($#,##0.00);"-"'
    ws.cell(tr, 11).font = header_font()
    ws.cell(tr, 11).alignment = right()


# ── Sheet 3: Performance ──────────────────────────────────────────────────────
def _build_performance(wb, df):
    ws = wb.create_sheet("📈 Performance")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Ticker",      12),
        ("Company",     28),
        ("Price",       13),
        ("1D Return",   13),
        ("1W Return",   13),
        ("1M Return",   13),
        ("1Y Return",   13),
        ("52W High",    13),
        ("52W Low",     13),
        ("vs 52W High", 14),
    ]

    ws.row_dimensions[1].height = 28
    for ci, (hdr, width) in enumerate(columns, 1):
        c = ws.cell(1, ci, hdr)
        c.font      = header_font()
        c.fill      = fill(DARK_NAVY)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    for i, row in df.iterrows():
        r  = i + 2
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        d1 = (row["Current Price"] - row["Prev Close"]) / row["Prev Close"]
        d7 = (row["Current Price"] - row["1W Ago"])     / row["1W Ago"]
        d30= (row["Current Price"] - row["1M Ago"])     / row["1M Ago"]
        d1y= (row["Current Price"] - row["1Y Ago"])     / row["1Y Ago"]
        hi = row["52W High"]
        lo = row["52W Low"]
        vs_high = (row["Current Price"] - hi) / hi if hi else 0

        data = [
            (row["Ticker"],           "str",   None),
            (row["Company"],          "str",   None),
            (row["Current Price"],    "price", None),
            (d1,                      "pct",   d1),
            (d7,                      "pct",   d7),
            (d30,                     "pct",   d30),
            (d1y,                     "pct",   d1y),
            (hi,                      "price", None),
            (lo,                      "price", None),
            (vs_high,                 "pct",   vs_high),
        ]

        for ci, (v, fmt, signal) in enumerate(data, 1):
            c = ws.cell(r, ci, v)
            c.fill   = fill(bg)
            c.border = thin_border()
            if fmt == "str":
                c.alignment = Alignment(vertical="center")
                c.font = body_font()
            elif fmt == "price":
                c.number_format = '$#,##0.00'
                c.alignment = right()
                c.font = body_font()
            elif fmt == "pct":
                c.number_format = '+0.00%;-0.00%;"-"'
                c.alignment = right()
                c.font = Font(name="Arial", size=10,
                              color=GREEN if (signal or 0) >= 0 else RED,
                              bold=abs(signal or 0) > 0.05)


# ── Sheet 4: Fundamentals ─────────────────────────────────────────────────────
def _build_fundamentals(wb, df):
    ws = wb.create_sheet("🔍 Fundamentals")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Ticker",     12),
        ("Company",    28),
        ("Sector",     22),
        ("Mkt Cap",    16),
        ("P/E Ratio",  13),
        ("Div Yield",  13),
        ("Beta",       10),
        ("52W High",   13),
        ("52W Low",    13),
        ("Price Range",14),
    ]

    ws.row_dimensions[1].height = 28
    for ci, (hdr, width) in enumerate(columns, 1):
        c = ws.cell(1, ci, hdr)
        c.font      = header_font()
        c.fill      = fill(DARK_NAVY)
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    for i, row in df.iterrows():
        r  = i + 2
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        hi = row["52W High"] or 0
        lo = row["52W Low"]  or 0
        rng = hi - lo

        data = [
            (row["Ticker"],                 "str"),
            (row["Company"],                "str"),
            (row["Sector"],                 "str"),
            (row["Market Cap"],             "mktcap"),
            (row["PE Ratio"],               "1dp"),
            (row["Div Yield"],              "pct"),
            (row["Beta"],                   "2dp"),
            (hi,                            "price"),
            (lo,                            "price"),
            (rng,                           "price"),
        ]

        for ci, (v, fmt) in enumerate(data, 1):
            val = v if v is not None else "N/A"
            c = ws.cell(r, ci, val)
            c.fill   = fill(bg)
            c.border = thin_border()
            c.font   = body_font()
            if fmt == "str":
                c.alignment = Alignment(vertical="center")
            elif fmt == "price":
                c.number_format = '$#,##0.00'
                c.alignment = right()
            elif fmt == "mktcap":
                c.number_format = '$#,##0,,,"B"'
                c.alignment = right()
            elif fmt == "pct":
                if isinstance(val, (int, float)):
                    c.number_format = '0.00%'
                c.alignment = right()
            elif fmt in ("1dp", "2dp"):
                if isinstance(val, (int, float)):
                    c.number_format = '0.0' if fmt == "1dp" else '0.00'
                c.alignment = right()


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("🔄  Fetching live market data...")
    df = fetch_data(PORTFOLIO)

    if df.empty:
        print("❌  No data retrieved. Check tickers and internet connection.")
        sys.exit(1)

    print(f"✅  Fetched data for {len(df)} holdings.")
    print("📊  Building Excel report...")
    build_workbook(df)
